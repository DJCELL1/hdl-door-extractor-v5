#!/usr/bin/env python3
"""
HDL Door Schedule Extractor v5
Author: Lianga (for Selwyn @ Hardware Direct)
Version: 5.0
"""
import os, re, pdfplumber, openpyxl, pytesseract
from pdf2image import convert_from_path
from PIL import Image
from openpyxl.styles import Font, PatternFill

HEADER_COLUMNS = ["Area", "Door", "Code", "Description", "Colour", "Quantity"]
SHEET_NAME = "Doors with Hardware"

def normalise_text(s):
    if not s: return ""
    s = s.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip()

def find_colour(desc):
    if not desc: return ""
    colours = ["SSS", "SC", "BLK", "BK", "BN", "AB", "PB", "PN", "SN", "MSB", "SIL", "SATIN", "CHROME", "BLACK", "WHITE", "BRASS"]
    for c in colours:
        if c in desc.upper(): return c
    return ""

def safe_quantity(q):
    if not q: return 1
    m = re.search(r"(\d{1,4})", q)
    return int(m.group(1)) if m else 1

class ExtractedLine:
    def __init__(self, area, door, code, description, colour, quantity, source="line"):
        self.area, self.door, self.code, self.description, self.colour, self.quantity, self.source = area, door, code, description, colour, quantity, source

class PDFExtractor:
    def __init__(self, path, dpi=300):
        self.path, self.dpi = path, dpi
    def text_pages(self):
        texts = []
        with pdfplumber.open(self.path) as pdf:
            for p in pdf.pages: texts.append(normalise_text(p.extract_text() or ""))
        return texts
    def ocr_pages(self):
        imgs = convert_from_path(self.path, dpi=self.dpi)
        texts = []
        for i in imgs:
            gray = i.convert("L").point(lambda x: 0 if x < 200 else 255, "1")
            texts.append(normalise_text(pytesseract.image_to_string(gray, config="--psm 6")))
        return texts

class ScheduleParser:
    def __init__(self, supplier="GENERIC"): self.supplier = supplier.upper()
    def parse_lines(self, pages):
        lines, line_re = [], re.compile(r"^(?P<area>[A-Za-z0-9 .\-/()]+?)\s+(?P<door>[A-Z]{1,4}\d{2,4}[A-Z]?(?:-[A-Z0-9])?)\s+(?P<code>[A-Za-z0-9./-]+)\s+(?P<qty>\d{1,4})?\s+(?P<desc>.+)$")
        for p in pages:
            for raw in p.splitlines():
                m = line_re.search(raw)
                if not m: continue
                area, door, code, desc = m.group("area"), m.group("door"), m.group("code"), m.group("desc")
                qty, colour = safe_quantity(m.group("qty")), find_colour(desc)
                lines.append(ExtractedLine(area, door, code, desc, colour, qty))
        return lines

def write_excel(lines, out_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = SHEET_NAME
    for i, h in enumerate(HEADER_COLUMNS, 1):
        c = ws.cell(1, i, h); c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9D9D9")
    for l in lines: ws.append([l.area, l.door, l.code, l.description, l.colour, l.quantity])
    wb.save(out_path)

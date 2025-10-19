# =============================================================================
# extractor.py | Core logic for HDL Door Schedule Extractor v5
# =============================================================================
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
import re
import tempfile
import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# -----------------------------------------------------------------------------
# Data structure for a parsed line
# -----------------------------------------------------------------------------
class ExtractedLine:
    def __init__(self, area, door, code, description, colour, quantity, source=None):
        self.area = area
        self.door = door
        self.code = code
        self.description = description
        self.colour = colour
        self.quantity = quantity
        self.source = source


# -----------------------------------------------------------------------------
# PDF Extractor: converts PDF pages to text (or OCR if needed)
# -----------------------------------------------------------------------------
class PDFExtractor:
    def __init__(self, path):
        self.path = path

    def text_pages(self):
        pages = []
        with pdfplumber.open(self.path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
        return pages

    def ocr_pages(self):
        images = convert_from_path(self.path, dpi=300)
        texts = []
        for img in images:
            text = pytesseract.image_to_string(img)
            texts.append(text)
        return texts


# -----------------------------------------------------------------------------
# Schedule Parser: parses raw text into structured lines
# -----------------------------------------------------------------------------
class ScheduleParser:
    def __init__(self, supplier):
        self.supplier = supplier.upper().strip()

    def parse_lines(self, pages):
        pattern = re.compile(r"(\b[A-Z]+\b.*)")  # basic fallback
        lines = []
        for text in pages:
            for raw_line in text.split("\n"):
                parts = [p.strip() for p in raw_line.split("\t") if p.strip()]
                if len(parts) >= 4:
                    area, door, code, *desc = parts
                    desc_text = " ".join(desc)
                    colour, qty = "", "1"
                    lines.append(ExtractedLine(area, door, code, desc_text, colour, qty))
        return lines


# -----------------------------------------------------------------------------
# Excel Writer: writes extracted lines to an Excel file
# -----------------------------------------------------------------------------
def write_excel(lines, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Doors with Hardware"
    header = ["Area", "Door", "Code", "Description", "Colour", "Quantity"]
    ws.append(header)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for line in lines:
        ws.append([line.area, line.door, line.code, line.description, line.colour, line.quantity])

    ws.freeze_panes = "A2"
    wb.save(output_path)
